from openpyxl.styles import (PatternFill, Border, Side, Alignment, Font)
from datetime import datetime, timedelta
from typing import NamedTuple

from brute_force_2.xlsx_generator.base_xlsx import Base
from config import DAYS


class ScheduleTime(NamedTuple):
    start_hour: int = 9
    start_minute: int = 0
    end_hour: int = 0
    end_minute: int = 0


class TableGenerator(Base):

    __slots__ = ('__title', '__sequence', '__work_time', '__time_increment', '__file_name', '__directory_name', '__sheet')

    def __init__(self, title, sequence, work_time=ScheduleTime(), time_increment=15):
        super().__init__(title)
        self.__title = f'Time table for {title}'
        self.__sequence = sequence
        self.__work_time = work_time
        self.__time_increment = timedelta(minutes=time_increment)
        self.__file_name = f'{self.__title}.xlsx'
        self.__sheet = self.get_work_sheet()

    def generate_table(self):
        self.set_base_template('A', 'F')
        self.generate_template()

        row_a = 0
        row_b = 0
        column = None
        


        for room_name in self.__sequence:
            ...
    def generate_template(self, mode: str = 'room'):
        self.set_time_column()
        columns = ('A', 'B', 'C', 'D', 'E', 'F')

        for i, cell in enumerate(columns):
            if cell == 'A':
                self.__sheet.column_dimensions['A'].width = 15
                self.write_text(column='A', row=3, text='Duration')
            else:
                self.__sheet.column_dimensions[cell].width = 35
                self.write_text(column=cell, row=3, text=DAYS[i - 1], font_type='general')

    def set_base_template(self, start: str, end: str):
        self.__sheet.merge_cells(f'{start}1:{end}1')
        self.__sheet.row_dimensions[1].height = 25
        self.color_cell(column='B', row=1, color='E0E0E0')
        self.color_cell(column='A', row=1, color='E0E0E0')
        self.write_text(column='B', row=1, text=self.__title, font_type='title')

    def set_time_column(self, row: int = 5, column: int = 1, start_time=None, end_time=None) -> None:
        if start_time is None:
            start_time = datetime(2023, 7, 12, self.__work_time.start_hour, self.__work_time.start_minute)
        if end_time is None:
            end_time = datetime(2023, 7, 13, self.__work_time.end_hour, self.__work_time.end_minute)
        if start_time > end_time:
            return
        self.write_text(row=row, column='A', text=start_time.strftime('%H:%M'), font_type='general')
        start_time += self.__time_increment
        self.set_time_column(row=row + 1, column=column, start_time=start_time, end_time=end_time)

    def text_location(self, column: str, row: int, wrap_text=True):
        self.__sheet[f'{column}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap_text)

    def font_text(self, column: str, row: int, font_type: str):
        locStyles: dict = {'general': {'name': 'Arial', 'size': 16, 'bold': False, 'color': '00000000'},
                           'general_bold': {'name': 'Arial', 'size': 16, 'bold': True, 'color': '00000000'},
                           'title': {'name': 'Arial', 'size': 22, 'bold': True, 'color': '00000000'},
                           'event': {'name': 'Arial', 'size': 16, 'bold': False, 'color': '00000000'},
                           'class': {'name': 'Arial', 'size': 14, 'bold': True, 'color': '00000000'},
                           'personal': {}, }
        self.__sheet[f'{column}{row}'].font = Font(**locStyles[font_type])

    def color_cell(self, column: str, row: int, color=None):
        if color is None:
            color = 'E0E0E0'
        self.__sheet[f'{column}{row}'].fill = PatternFill(fill_type='solid', start_color=color, end_color=color, )

    def borderStyle(self, column: str, row: int, border_type: str):
        loc_border = Side(border_style=border_type)
        self.__sheet[f'{column}{row}'].border = Border(left=loc_border, right=loc_border, bottom=loc_border, top=loc_border)

    def write_text(self, column: str, row: int, text: str, font_type: str = 'general', wrap_text: bool = True):
        self.__sheet[f'{column}{row}'] = text
        self.text_location(column=column, row=row, wrap_text=wrap_text)
        self.font_text(column=column, row=row, font_type=font_type)

    def borderSet(self, column: int, start_row: int, end_row: int):
        loc_border = Side(border_style='medium')
        for row in self.__sheet.iter_rows(min_row=start_row, min_col=column, max_row=end_row, max_col=column):
            for cell in row:
                cell.border = Border(left=loc_border, right=loc_border, bottom=loc_border, top=loc_border)

