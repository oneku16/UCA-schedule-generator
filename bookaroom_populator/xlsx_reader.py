from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from json import loads, dump

INDEXES = {'day': 0,
           'cohort': 1}


class XlsxToJSON:
    __slots__ = ('__xlsx',)

    def __init__(self, xlsx: Worksheet):
        self.__xlsx = xlsx

    def run(self):
        for cell in self.__xlsx:
            print(cell)


def get_xlsx_object(file_name: str) -> Workbook:
    try:
        with open(file_name, 'rb') as xlsx:
            return load_workbook(xlsx)
    except BaseException:
        raise 'Incorrect xlsx file'


def main():
    xlsx = get_xlsx_object('target.xlsx')
    sheet: Worksheet = xlsx[next(iter(xlsx.sheetnames))]
    base = XlsxToJSON(sheet)
    base.run()


if __name__ == '__main__':
    main()
