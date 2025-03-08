from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from os import path, walk, mkdir

from src.schedule.consts import BASE_DIR


dir_path = path.dirname(path.realpath(__file__))


class Base:
    __slots__ = ('_file_name', '_file_path', '_folder_dir', '_work_book')

    def __init__(self, filename: str):
        self._file_name = filename
        self._file_path = path.join(BASE_DIR, 'xlsx_files', filename)
        self._folder_dir = path.join(BASE_DIR, 'xlsx_files')
        self._work_book = None

    def get_xlsx_from_directory(self) -> set[str]:
        files = set()
        if not path.isdir(self._folder_dir):
            mkdir(self._folder_dir)
            return files

        for (_, _, d_files) in walk(self._folder_dir):
            for file in d_files:
                files.add(file)

        return files

    def get_work_sheet(self) -> Worksheet:

        if self._file_name in self.get_xlsx_from_directory():
            with open(self._file_path, 'rb') as xlsx:
                self._work_book = load_workbook(xlsx)
                work_sheet = self._work_book[self._work_book.sheetnames[0]]
                return work_sheet

        self._work_book = Workbook()
        work_sheet = self._work_book.active
        return work_sheet

    def normalize_file_name(self):
        if self._file_name.endswith(('.xlsx', 'exel')):
            return path.join(self._folder_dir, self._file_name)
        return path.join(self._folder_dir, f'{self._file_name}.xlsx')

    def save_xlsx(self):
        self._work_book.save(filename=self.normalize_file_name())
        self._work_book.close()

    @property
    def file_path(self):
        return self._file_path

    @property
    def file_name(self):
        return self._file_name

    @property
    def folder_dir(self):
        return self._folder_dir

    @property
    def work_book(self):
        return self._work_book
