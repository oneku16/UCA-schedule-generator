from openpyxl import load_workbook
from config import DEANS_MEMO_PATH, DEPARTMENT_NAMES, SUBJECT_JSON
from typing import List, Any, Dict, Optional, Union
from copy import deepcopy


class Converter:
    def __init__(self):
        self.xlsx_file = load_workbook(DEANS_MEMO_PATH, data_only=True)
        self.tbd_index = 0

    def get_xlsx_sheet(self, sheet_name: Union[int, str]) -> Any:
        if isinstance(sheet_name, int):
            try:
                return self.xlsx_file[self.xlsx_file.sheetnames[sheet_name]]
            except IndexError:
                raise ValueError("Sheet index out of range.")
        else:
            try:
                return self.xlsx_file[sheet_name]
            except KeyError:
                raise ValueError("Sheet name does not exist.")

    @staticmethod
    def get_subject_id(subject_area: Any, course_code: Any) -> str:
        return f'{str(subject_area).strip()}{str(course_code).strip()}'.strip()

    @staticmethod
    def get_subject_title(subject_title: str) -> str:
        return subject_title.replace(u'\xa0', u' ').strip()

    @staticmethod
    def get_subject_patterns(course_types: str) -> List[Dict[str, int]]:
        def _splitter():
            try:
                for subject_pattern in course_types.split(','):
                    subject_pattern = list(int(item.strip()) for item in subject_pattern.split('x'))
                    yield {'classes': subject_pattern[0], 'duration': subject_pattern[1]}
            except AttributeError:
                return [course_types]

        return list(_splitter())

    @staticmethod
    def distribute_by_cohort(year_level: str, cohort: str, cohort_number=None) -> str:
        if cohort_number:
            return f'Group {cohort_number[0]} {cohort}'.strip()
        return f'{"".join(map(str, filter(str.isupper, year_level)))} {cohort}'.strip()

    def get_instructor_names(self, primary_instructor: str, secondary_instructor: Optional[str]) -> Dict[str, Dict[str, str or None]]:
        instructor_info = {'primary': {'instructor_id': None, 'instructor_name': self.get_next_tbd_index(), 'preferences': None}}

        if primary_instructor and not primary_instructor.strip() == 'TBD':
            instructor_info = {'primary': {'instructor_id': None, 'instructor_name': primary_instructor.strip(), 'preferences': None}}
            if secondary_instructor:
                instructor_info['secondary'] = {'instructor_id': None, 'instructor_name': secondary_instructor.strip(), 'preferences': None}

        return instructor_info

    def get_next_tbd_index(self) -> str:
        self.tbd_index += 1
        return f'TBD{self.tbd_index}'

    def xlsx_to_json(self, sheet_name='Spring 2023') -> List[Dict]:
        def _producer():
            sheet = self.get_xlsx_sheet(sheet_name)
            cohort = None
            for row in sheet.iter_rows(min_row=2, values_only=True):

                subject_json = deepcopy(SUBJECT_JSON)

                if row[0] in DEPARTMENT_NAMES:
                    cohort = row[0]
                    continue

                subject_id = self.get_subject_id(row[0], row[1])
                subject_title = self.get_subject_title(row[2])
                cohort_distribution = self.distribute_by_cohort(cohort, row[6], row[8])
                subject_patterns = self.get_subject_patterns(row[11])
                instructors = self.get_instructor_names(row[12], row[13])

                if subject_id:
                    subject_json['id'] = subject_id
                if subject_title:
                    subject_json['title'] = subject_title
                if cohort_distribution:
                    subject_json['cohort'] = cohort_distribution
                if subject_patterns:
                    subject_json['patterns'] = subject_patterns
                if instructors:
                    subject_json['instructors'] = instructors

                yield subject_json

        return list(_producer())
