from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Font,
)
from datetime import datetime, timedelta
from typing import NamedTuple

from brute_force_3.xlsx_generator.base_xlsx import Base
from config import DAYS, ROOM_COLOR


class ScheduleTime(NamedTuple):
    start_hour: int = 9
    start_minute: int = 0
    end_hour: int = 0
    end_minute: int = 0


class TableGenerator(Base):
    __slots__ = (
            '__title', '__sequence', '__work_time', '__time_increment', '__file_name', '__directory_name', '__sheet')

    def __init__(self, title, sequence, work_time=ScheduleTime(), time_increment=15):
        super().__init__(title)
        self.__title = f'Time table for {title}'
        self.__sequence = sequence
        self.__work_time = work_time
        self.__time_increment = timedelta(minutes=time_increment)
        self.__file_name = f'{self.__title}.xlsx'
        self.__sheet = self.get_work_sheet()

    def generate_table(self, mode='room'):
        self.set_base_template('B', 'G')
        self.generate_template()
        columns = ('A', 'B', 'C', 'D', 'E', 'F')

        for values in self.__sequence:
            row_a = row_b = column = None
            start_time, end_time = values['start_time'], values['end_time']
            subject = values['subject']
            room = values['room']
            day = values['day']

            for row_index, cell in enumerate(self.__sheet['A']):
                if cell.value == start_time:
                    row_a = row_index
                    continue
                if cell.value == end_time:
                    row_b = row_index
                    break

            for column_index in range(1, len(columns)):
                if self.__sheet[f'{columns[column_index]}3'].value == day:
                    column = column_index
                    break

            cell_value: str = f'{subject}\n\n{start_time}-{end_time}\nroom:{room}'

            self.border_set(column=column + 1, start_row=row_a + 1, end_row=row_b)
            self.color_cell(column=columns[column], row=row_a + 1, color=ROOM_COLOR.get(room.split()[-1], 'FFFFFF'))
            self.__sheet.merge_cells(f'{columns[column]}{row_a + 1}:{columns[column]}{row_b}')
            self.write_text(column=columns[column], row=row_a + 1, text=cell_value, font_type='class')

        self.save_xlsx()

    def generate_template(self):
        self.set_time_column()
        columns = ('A', 'B', 'C', 'D', 'E', 'F')

        for i, cell in enumerate(columns):
            self.__sheet.merge_cells(f'{columns[i]}3:{columns[i]}4')
            self.color_cell(column=columns[i], row=3)
            if cell == 'A':
                self.__sheet.column_dimensions['A'].width = 15
                self.write_text(column='A', row=3, text='Duration')
            else:
                self.__sheet.column_dimensions[cell].width = 35
                self.write_text(column=cell, row=3, text=DAYS[i - 1], font_type='general')

    def set_base_template(self, start: str, end: str):
        self.__sheet.merge_cells(f'{start}1:{end}1')
        self.__sheet.row_dimensions[1].height = 25
        self.color_cell(column='B', row=1, color='E0E0E0')
        self.color_cell(column='A', row=1, color='E0E0E0')
        self.write_text(column='B', row=1, text=self.__title, font_type='title')

    def set_time_column(self, row: int = 6, column: int = 1, start_time=None, end_time=None) -> None:
        if start_time is None:
            start_time = datetime(2023, 7, 12, self.__work_time.start_hour, self.__work_time.start_minute)
        if end_time is None:
            end_time = datetime(2023, 7, 13, self.__work_time.end_hour, self.__work_time.end_minute)
        if start_time > end_time:
            return
        self.write_text(row=row, column='A', text=start_time.strftime('%H:%M'), font_type='general')
        start_time += self.__time_increment
        self.set_time_column(row=row + 1, column=column, start_time=start_time, end_time=end_time)

    def text_location(self, column: str, row: int, wrap_text=True):
        self.__sheet[f'{column}{row}'].alignment = Alignment(horizontal='center', vertical='center',
                                                             wrap_text=wrap_text)

    def font_text(self, column: str, row: int, font_type: str):
        locStyles: dict = {'general': {'name': 'Arial', 'size': 16, 'bold': False, 'color': '00000000'},
                           'general_bold': {'name': 'Arial', 'size': 16, 'bold': True, 'color': '00000000'},
                           'title': {'name': 'Arial', 'size': 22, 'bold': True, 'color': '00000000'},
                           'event': {'name': 'Arial', 'size': 16, 'bold': False, 'color': '00000000'},
                           'class': {'name': 'Arial', 'size': 14, 'bold': True, 'color': '00000000'},
                           'personal': {}, }
        self.__sheet[f'{column}{row}'].font = Font(**locStyles[font_type])

    def color_cell(self, column: str, row: int, color=None):
        if color is None:
            color = 'E0E0E0'
        self.__sheet[f'{column}{row}'].fill = PatternFill(fill_type='solid',
                                                          start_color=color,
                                                          end_color=color)

    def border_style(self, column: str, row: int, border_type: str):
        loc_border = Side(border_style=border_type)
        self.__sheet[f'{column}{row}'].border = Border(left=loc_border,
                                                       right=loc_border,
                                                       bottom=loc_border,
                                                       top=loc_border)

    def border_set(self, column: int, start_row: int, end_row: int):
        loc_border = Side(border_style='medium')
        for row in self.__sheet.iter_rows(min_row=start_row,
                                          min_col=column,
                                          max_row=end_row,
                                          max_col=column):
            for cell in row:
                cell.border = Border(left=loc_border,
                                     right=loc_border,
                                     bottom=loc_border,
                                     top=loc_border)

    def write_text(self, column: str, row: int, text: str, font_type: str = 'general', wrap_text: bool = True):
        self.__sheet[f'{column}{row}'] = text
        self.text_location(column=column,
                           row=row,
                           wrap_text=wrap_text)
        self.font_text(column=column,
                       row=row,
                       font_type=font_type)
